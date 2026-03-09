"""
saas_export.py — SRP MediFlow Data Export Module
=================================================
Generates hospital operational data exports for Admin users.

Formats  : Excel (.xlsx), CSV (.csv), PDF (.pdf)
Ranges   : daily | weekly | monthly | yearly | custom (from_date / to_date)
Exports  : patients | billing | appointments

Requirements (install once):
    pip install openpyxl reportlab

Usage (called from server handler):
    from saas_export import export_data
    file_bytes, mime, filename = export_data(
        export_type='patients',
        fmt='excel',
        date_range='monthly',
    )
"""

from __future__ import annotations
import csv
import io
import os
from datetime import datetime, timedelta, date
from typing import Optional

# ── Date range helpers ────────────────────────────────────────────────────────
def _date_range(
    range_key: str,
    from_date: Optional[str] = None,
    to_date:   Optional[str] = None,
) -> tuple[date, date]:
    """
    Return (start_date, end_date) based on range_key.
    range_key: 'daily' | 'weekly' | 'monthly' | 'yearly' | 'custom'
    For 'custom', from_date and to_date (ISO 8601) must be provided.
    """
    today = date.today()
    if range_key == "daily":
        return today, today
    elif range_key == "weekly":
        return today - timedelta(days=6), today
    elif range_key == "monthly":
        return today.replace(day=1), today
    elif range_key == "yearly":
        return today.replace(month=1, day=1), today
    elif range_key == "custom":
        try:
            start = date.fromisoformat(from_date) if from_date else today - timedelta(days=30)
            end   = date.fromisoformat(to_date)   if to_date   else today
            return start, end
        except Exception:
            return today - timedelta(days=30), today
    else:
        return today - timedelta(days=30), today


# ── DB fetch helpers ──────────────────────────────────────────────────────────
def _get_conn():
    import db
    return db.get_connection()


def _fetch_patients(start: date, end: date) -> list[dict]:
    conn = _get_conn()
    if not conn:
        return []
    try:
        import psycopg2.extras
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT id, name, age, phone, aadhar, issue, doctor,
                   appointment_time, status, source, created_at
              FROM registrations
             WHERE created_at::date BETWEEN %s AND %s
             ORDER BY created_at DESC
        """, (start, end))
        rows = [dict(r) for r in cur.fetchall()]
        cur.close(); conn.close()
        return rows
    except Exception as exc:
        print(f"[export] _fetch_patients error: {exc}")
        try: conn.close()
        except: pass
        return []


def _fetch_billing(start: date, end: date) -> list[dict]:
    conn = _get_conn()
    if not conn:
        return []
    try:
        import psycopg2.extras
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT id, patient_name, patient_phone, bill_type,
                   consultation_fee, lab_charges, pharmacy_charges,
                   bed_charges, surgery_charges, total_amount,
                   tax_amount, discount, net_amount, status,
                   created_by, created_at
              FROM billing
             WHERE created_at::date BETWEEN %s AND %s
             ORDER BY created_at DESC
        """, (start, end))
        rows = [dict(r) for r in cur.fetchall()]
        cur.close(); conn.close()
        return rows
    except Exception as exc:
        print(f"[export] _fetch_billing error: {exc}")
        try: conn.close()
        except: pass
        return []


def _fetch_appointments(start: date, end: date) -> list[dict]:
    conn = _get_conn()
    if not conn:
        return []
    try:
        import psycopg2.extras
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT id, patient_name, patient_phone, age, issue,
                   doctor_name, department, appointment_date,
                   appointment_time, status, source, created_at
              FROM appointments
             WHERE created_at::date BETWEEN %s AND %s
             ORDER BY created_at DESC
        """, (start, end))
        rows = [dict(r) for r in cur.fetchall()]
        cur.close(); conn.close()
        return rows
    except Exception as exc:
        print(f"[export] _fetch_appointments error: {exc}")
        try: conn.close()
        except: pass
        return []


# ── Format renderers ──────────────────────────────────────────────────────────

def _to_csv(rows: list[dict]) -> bytes:
    if not rows:
        return b"No data available for selected range\n"
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    for row in rows:
        # Convert datetimes to strings
        clean = {k: (str(v) if not isinstance(v, (str, int, float, type(None))) else v)
                 for k, v in row.items()}
        writer.writerow(clean)
    return buf.getvalue().encode("utf-8-sig")   # BOM for Excel compatibility


def _to_excel(rows: list[dict], sheet_title: str = "Export") -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        # Fallback to CSV if openpyxl not installed
        return _to_csv(rows)

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = sheet_title[:31]   # Excel sheet name limit

    if not rows:
        ws["A1"] = "No data available for selected range"
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # Header styling
    header_fill = PatternFill("solid", fgColor="1A73E8")
    header_font = Font(bold=True, color="FFFFFF")

    headers = list(rows[0].keys())
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header.replace("_", " ").title())
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(headers, 1):
            val = row.get(key)
            if hasattr(val, "isoformat"):
                val = val.isoformat()
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # Summary row
    summary_row = ws.max_row + 2
    ws.cell(row=summary_row, column=1, value=f"Total records: {len(rows)}")
    ws.cell(row=summary_row, column=1).font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _to_pdf(rows: list[dict], title: str = "SRP MediFlow Export") -> bytes:
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
    except ImportError:
        # Fallback to CSV if reportlab not installed
        return _to_csv(rows)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30
    )
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph(f"<b>{title}</b>", styles["Title"]))
    elements.append(Paragraph(
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  Records: {len(rows)}",
        styles["Normal"]
    ))
    elements.append(Spacer(1, 12))

    if not rows:
        elements.append(Paragraph("No data available for selected range.", styles["Normal"]))
    else:
        headers = list(rows[0].keys())
        # Limit columns to keep PDF readable
        display_cols = headers[:8]
        table_data   = [[h.replace("_", " ").title() for h in display_cols]]
        for row in rows[:500]:      # cap at 500 rows for PDF
            table_data.append([
                str(row.get(h, "") or "")[:40]
                for h in display_cols
            ])

        tbl = Table(table_data, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0),  colors.HexColor("#1A73E8")),
            ("TEXTCOLOR",   (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, -1), 8),
            ("GRID",        (0, 0), (-1, -1), 0.4, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F9FA")]),
            ("ALIGN",       (0, 0), (-1, -1), "LEFT"),
            ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING",  (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(tbl)

    doc.build(elements)
    return buf.getvalue()


# ── Public API ────────────────────────────────────────────────────────────────

# Map export_type → (fetch fn, sheet title, PDF title)
_EXPORTERS = {
    "patients":     (_fetch_patients,     "Patients",     "Patient Registrations"),
    "billing":      (_fetch_billing,      "Billing",      "Billing Report"),
    "appointments": (_fetch_appointments, "Appointments", "Appointments Report"),
}

_MIME = {
    "excel": ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", ".xlsx"),
    "csv":   ("text/csv; charset=utf-8",                                            ".csv"),
    "pdf":   ("application/pdf",                                                    ".pdf"),
}


def export_data(
    export_type: str,
    fmt:         str   = "excel",
    date_range:  str   = "monthly",
    from_date:   Optional[str] = None,
    to_date:     Optional[str] = None,
) -> tuple[bytes, str, str]:
    """
    Generate export file bytes.

    Parameters
    ----------
    export_type : 'patients' | 'billing' | 'appointments'
    fmt         : 'excel'    | 'csv'     | 'pdf'
    date_range  : 'daily' | 'weekly' | 'monthly' | 'yearly' | 'custom'
    from_date   : ISO date string (only used when date_range='custom')
    to_date     : ISO date string (only used when date_range='custom')

    Returns
    -------
    (file_bytes, mime_type, filename)
    """
    fetch_fn, sheet_title, pdf_title = _EXPORTERS.get(
        export_type.lower(),
        _EXPORTERS["patients"]
    )
    mime_type, ext = _MIME.get(fmt.lower(), _MIME["csv"])

    start, end  = _date_range(date_range, from_date, to_date)
    rows        = fetch_fn(start, end)
    ts          = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename    = f"srp_{export_type}_{date_range}_{ts}{ext}"

    if fmt.lower() == "excel":
        file_bytes = _to_excel(rows, sheet_title)
    elif fmt.lower() == "pdf":
        file_bytes = _to_pdf(rows, f"{pdf_title} — {start} to {end}")
    else:
        file_bytes = _to_csv(rows)

    return file_bytes, mime_type, filename
